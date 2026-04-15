from __future__ import annotations

from pathlib import Path
from typing import Literal
from threading import Thread, RLock
from datetime import datetime
import time
import re
import unicodedata
import os

import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "dados"
DATA_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = DATA_DIR / "dados_fluxo_caixa.xlsx"

SHEETS = {
    "Entradas": ["id", "data", "descricao", "categoria", "valor", "status"],
    "Saidas": ["id", "data", "descricao", "categoria", "forma_pagamento", "valor", "status"],
    "DespesasMes": ["id", "conta_mes", "descricao", "vencimento", "forma_pagamento", "valor", "status"],
}

# =========================
# VARIÁVEIS DE AMBIENTE
# =========================
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_ALLOWED_CHAT_ID = os.getenv("TELEGRAM_ALLOWED_CHAT_ID", "").strip()

# Exemplo:
# BACKEND_CORS_ORIGINS=http://localhost:5173,http://127.0.0.1:5173,https://SEU_USUARIO.github.io
cors_origins_env = os.getenv(
    "BACKEND_CORS_ORIGINS",
    "http://localhost:5173,http://127.0.0.1:5173"
).strip()

ALLOWED_ORIGINS = [
    origin.strip()
    for origin in cors_origins_env.split(",")
    if origin.strip()
]

TELEGRAM_API_URL = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}" if TELEGRAM_BOT_TOKEN else ""
telegram_offset = 0
workbook_lock = RLock()


class EntradaIn(BaseModel):
    data: str
    descricao: str = Field(min_length=1)
    categoria: str = Field(min_length=1)
    valor: float = Field(ge=0)
    status: str = "Recebido"


class EntradaOut(EntradaIn):
    id: int


class SaidaIn(BaseModel):
    data: str
    descricao: str = Field(min_length=1)
    categoria: str = Field(min_length=1)
    forma_pagamento: str = Field(min_length=1)
    valor: float = Field(ge=0)
    status: str = "Pago"


class SaidaOut(SaidaIn):
    id: int


class DespesaIn(BaseModel):
    conta_mes: str = Field(min_length=1)
    descricao: str = Field(min_length=1)
    vencimento: str
    forma_pagamento: str = Field(min_length=1)
    valor: float = Field(ge=0)
    status: Literal["Pendente", "Pago", "Vencido"] = "Pendente"


class DespesaOut(DespesaIn):
    id: int


app = FastAPI(title="Fluxo de Caixa API", version="1.3.0")

# IMPORTANTE:
# se usar allow_credentials=True, não use allow_origins=["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS if ALLOWED_ORIGINS else ["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def init_workbook() -> None:
    with workbook_lock:
        if XLSX_PATH.exists():
            wb = load_workbook(XLSX_PATH)
            changed = False

            for sheet_name, headers in SHEETS.items():
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(headers)
                    changed = True

            if changed:
                wb.save(XLSX_PATH)

            wb.close()
            return

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        for sheet_name, headers in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)

        wb.save(XLSX_PATH)
        wb.close()


def get_workbook():
    init_workbook()
    return load_workbook(XLSX_PATH)


def rows_as_dicts(sheet_name: str) -> list[dict]:
    with workbook_lock:
        wb = get_workbook()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    if not rows:
        return []

    headers = list(rows[0])
    data: list[dict] = []

    for row in rows[1:]:
        if row and any(cell is not None for cell in row):
            item = {headers[idx]: row[idx] for idx in range(len(headers))}
            if item.get("id") is not None:
                item["id"] = int(item["id"])
            if item.get("valor") is not None:
                item["valor"] = float(item["valor"])
            data.append(item)

    return data


def next_id(sheet_name: str) -> int:
    items = rows_as_dicts(sheet_name)
    if not items:
        return 1
    return max(int(item["id"]) for item in items) + 1


def append_row(sheet_name: str, row: dict) -> dict:
    with workbook_lock:
        print(f"[Excel] append_row iniciado em {sheet_name}")
        wb = get_workbook()
        ws = wb[sheet_name]
        headers = SHEETS[sheet_name]
        row_id = next_id(sheet_name)
        payload = {"id": row_id, **row}

        ws.append([payload.get(col) for col in headers])
        wb.save(XLSX_PATH)
        wb.close()

        print(f"[Excel] append_row finalizado em {sheet_name}: {payload}")
        return payload


def update_row(sheet_name: str, item_id: int, row: dict) -> dict:
    with workbook_lock:
        wb = get_workbook()
        ws = wb[sheet_name]
        headers = SHEETS[sheet_name]

        for excel_row in range(2, ws.max_row + 1):
            current_id = ws.cell(excel_row, 1).value
            if current_id == item_id:
                payload = {"id": item_id, **row}
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(excel_row, col_idx).value = payload.get(header)
                wb.save(XLSX_PATH)
                wb.close()
                return payload

        wb.close()

    raise HTTPException(status_code=404, detail=f"Item {item_id} não encontrado em {sheet_name}.")


def delete_row(sheet_name: str, item_id: int) -> dict:
    with workbook_lock:
        wb = get_workbook()
        ws = wb[sheet_name]

        for excel_row in range(2, ws.max_row + 1):
            current_id = ws.cell(excel_row, 1).value
            if current_id == item_id:
                ws.delete_rows(excel_row, 1)
                wb.save(XLSX_PATH)
                wb.close()
                return {"success": True, "id": item_id}

        wb.close()

    raise HTTPException(status_code=404, detail=f"Item {item_id} não encontrado em {sheet_name}.")


def format_currency(value: float) -> str:
    inteiro, decimal = f"{value:.2f}".split(".")
    inteiro = f"{int(inteiro):,}".replace(",", ".")
    return f"R$ {inteiro},{decimal}"


def parse_valor(texto: str) -> float:
    texto = texto.strip().replace("R$", "").replace(" ", "")
    texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def hoje_br() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def normalize_text(texto: str) -> str:
    texto = (texto or "").strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def calcular_resumo() -> dict:
    entradas = rows_as_dicts("Entradas")
    saidas = rows_as_dicts("Saidas")
    despesas = rows_as_dicts("DespesasMes")

    total_entradas = sum(float(item.get("valor", 0) or 0) for item in entradas)

    total_saidas_base = sum(
        float(item.get("valor", 0) or 0)
        for item in saidas
        if str(item.get("forma_pagamento", "")).strip().lower() not in ["cartão de crédito", "cartao de credito"]
    )

    total_despesas_pagas = sum(
        float(item.get("valor", 0) or 0)
        for item in despesas
        if str(item.get("status", "")).strip() == "Pago"
        and str(item.get("forma_pagamento", "")).strip().lower() not in ["cartão de crédito", "cartao de credito"]
    )

    total_cartao = sum(
        float(item.get("valor", 0) or 0)
        for item in saidas
        if str(item.get("forma_pagamento", "")).strip().lower() in ["cartão de crédito", "cartao de credito"]
    )

    saldo_atual = total_entradas - (total_saidas_base + total_despesas_pagas)

    return {
        "saldo_atual": saldo_atual,
        "total_entradas": total_entradas,
        "total_saidas": total_saidas_base + total_despesas_pagas,
        "total_cartao_credito": total_cartao,
    }


def telegram_send_message(chat_id: str, text: str) -> None:
    if not TELEGRAM_API_URL:
        return

    try:
        response = requests.post(
            f"{TELEGRAM_API_URL}/sendMessage",
            json={"chat_id": chat_id, "text": text},
            timeout=20,
        )
        print(f"[Telegram] sendMessage => {response.status_code} | {response.text}")
    except Exception as exc:
        print(f"[Telegram] Erro ao enviar mensagem: {exc}")


def telegram_delete_webhook():
    if not TELEGRAM_API_URL:
        return

    try:
        r = requests.get(f"{TELEGRAM_API_URL}/deleteWebhook", timeout=20)
        print(f"[Telegram] deleteWebhook => {r.status_code} | {r.text}")
    except Exception as exc:
        print(f"[Telegram] Erro ao remover webhook: {exc}")


def telegram_get_me():
    if not TELEGRAM_API_URL:
        return

    try:
        r = requests.get(f"{TELEGRAM_API_URL}/getMe", timeout=20)
        print(f"[Telegram] getMe => {r.status_code} | {r.text}")
    except Exception as exc:
        print(f"[Telegram] Erro no getMe: {exc}")


def telegram_get_updates(offset: int = 0) -> list[dict]:
    if not TELEGRAM_API_URL:
        return []

    try:
        response = requests.get(
            f"{TELEGRAM_API_URL}/getUpdates",
            params={
                "timeout": 25,
                "offset": offset,
            },
            timeout=35,
        )
        response.raise_for_status()
        data = response.json()
        return data.get("result", [])
    except Exception as exc:
        print(f"[Telegram] Erro ao consultar updates: {exc}")
        return []


def interpretar_mensagem(texto: str) -> dict:
    texto_original = (texto or "").strip()

    if not texto_original:
        return {"tipo": "ajuda"}

    texto = normalize_text(texto_original)

    if texto in ["/start", "start", "/help", "help", "ajuda"]:
        return {"tipo": "ajuda"}

    if texto in ["/saldo", "saldo"]:
        return {"tipo": "saldo"}

    if texto in ["/resumo", "resumo"]:
        return {"tipo": "resumo"}

    m = re.match(r"^entrada\s+([0-9\.,]+)\s+(.+)$", texto, flags=re.IGNORECASE)
    if m:
        valor = parse_valor(m.group(1))
        resto = m.group(2).strip()

        partes = resto.split(" ", 1)
        categoria = partes[0].title()
        descricao = partes[1].strip() if len(partes) > 1 else categoria

        return {
            "tipo": "entrada",
            "payload": {
                "data": hoje_br(),
                "descricao": descricao,
                "categoria": categoria,
                "valor": valor,
                "status": "Recebido",
            },
        }

    m = re.match(r"^saida\s+([0-9\.,]+)\s+(.+)$", texto, flags=re.IGNORECASE)
    if m:
        valor = parse_valor(m.group(1))
        resto = m.group(2).strip()

        formas = [
            ("cartao de credito", "Cartão de Crédito"),
            ("cartao de debito", "Cartão de Débito"),
            ("transferencia", "Transferência"),
            ("boleto", "Boleto"),
            ("pix", "PIX"),
            ("dinheiro", "Dinheiro"),
        ]

        forma_pagamento = "PIX"
        descricao = resto

        for chave, valor_forma in formas:
            if resto.startswith(chave):
                forma_pagamento = valor_forma
                descricao = resto[len(chave):].strip()
                break

        descricao = descricao or "Lançamento via Telegram"
        categoria = descricao.split(" ")[0].title() if descricao else "Outros"

        return {
            "tipo": "saida",
            "payload": {
                "data": hoje_br(),
                "descricao": descricao,
                "categoria": categoria,
                "forma_pagamento": forma_pagamento,
                "valor": valor,
                "status": "Pago",
            },
        }

    m = re.match(r"^despesas?\s+([0-9\.,]+)\s+(.+)$", texto, flags=re.IGNORECASE)
    if m:
        valor = parse_valor(m.group(1))
        resto = m.group(2).strip()

        formas = [
            ("transferencia", "Transferência"),
            ("boleto", "Boleto"),
            ("pix", "PIX"),
            ("dinheiro", "Dinheiro"),
        ]

        forma_pagamento = "PIX"
        descricao = resto

        for chave, valor_forma in formas:
            if resto.startswith(chave):
                forma_pagamento = valor_forma
                descricao = resto[len(chave):].strip()
                break

        status = "Pendente"
        vencimento = hoje_br()

        status_match = re.search(r"\b(pago|pendente|vencido)$", descricao)
        if status_match:
            status_txt = status_match.group(1)
            descricao = descricao[:status_match.start()].strip()
            if status_txt == "pago":
                status = "Pago"
            elif status_txt == "vencido":
                status = "Vencido"
            else:
                status = "Pendente"

        data_match = re.search(r"(\d{2}/\d{2}/\d{4})$", descricao)
        if data_match:
            vencimento = data_match.group(1)
            descricao = descricao[:data_match.start()].strip()

        descricao = descricao or "Despesa via Telegram"
        conta_mes = descricao.split(" ")[0].title() if descricao else "Outros"

        return {
            "tipo": "despesa",
            "payload": {
                "conta_mes": conta_mes,
                "descricao": descricao,
                "vencimento": vencimento,
                "forma_pagamento": forma_pagamento,
                "valor": valor,
                "status": status,
            },
        }

    return {"tipo": "ajuda"}


def processar_update(update: dict) -> None:
    message = update.get("message") or {}
    chat = message.get("chat") or {}
    chat_id = str(chat.get("id", "")).strip()
    chat_type = str(chat.get("type", "")).strip()
    first_name = message.get("from", {}).get("first_name", "")
    username = message.get("from", {}).get("username", "")
    texto = message.get("text", "")

    if not chat_id:
        return

    print("=" * 90)
    print("[Telegram] Nova mensagem recebida")
    print(f"[Telegram] chat_id   : {chat_id}")
    print(f"[Telegram] chat_type : {chat_type}")
    print(f"[Telegram] first_name: {first_name}")
    print(f"[Telegram] username  : {username}")
    print(f"[Telegram] texto     : {texto}")
    print("=" * 90)

    if TELEGRAM_ALLOWED_CHAT_ID and chat_id != TELEGRAM_ALLOWED_CHAT_ID:
        telegram_send_message(chat_id, "Este chat não está autorizado para lançar dados.")
        return

    acao = interpretar_mensagem(texto)
    print(f"[Telegram] Ação interpretada: {acao}")

    try:
        if acao["tipo"] == "entrada":
            print(f"[Telegram] Processando entrada. Payload: {acao['payload']}")
            item = append_row("Entradas", acao["payload"])
            print(f"[Telegram] Entrada salva no Excel: {item}")

            telegram_send_message(
                chat_id,
                "✅ Entrada registrada com sucesso\n\n"
                f"ID: {item['id']}\n"
                f"Data: {item['data']}\n"
                f"Categoria: {item['categoria']}\n"
                f"Descrição: {item['descricao']}\n"
                f"Valor: {format_currency(float(item['valor']))}"
            )
            print(f"[Telegram] Resposta enviada ao chat_id={chat_id}")
            return

        if acao["tipo"] == "saida":
            print(f"[Telegram] Processando saída. Payload: {acao['payload']}")
            item = append_row("Saidas", acao["payload"])
            print(f"[Telegram] Saída salva no Excel: {item}")

            aviso = ""
            if str(item["forma_pagamento"]).strip().lower() in ["cartão de crédito", "cartao de credito"]:
                aviso = "\n\nℹ️ Registrado apenas como cartão de crédito. Não debita do saldo atual."

            telegram_send_message(
                chat_id,
                "✅ Saída registrada com sucesso\n\n"
                f"ID: {item['id']}\n"
                f"Data: {item['data']}\n"
                f"Forma: {item['forma_pagamento']}\n"
                f"Descrição: {item['descricao']}\n"
                f"Valor: {format_currency(float(item['valor']))}"
                f"{aviso}"
            )
            print(f"[Telegram] Resposta enviada ao chat_id={chat_id}")
            return

        if acao["tipo"] == "despesa":
            print(f"[Telegram] Processando despesa. Payload: {acao['payload']}")
            item = append_row("DespesasMes", acao["payload"])
            print(f"[Telegram] Despesa salva no Excel: {item}")

            telegram_send_message(
                chat_id,
                "✅ Despesa registrada com sucesso\n\n"
                f"ID: {item['id']}\n"
                f"Vencimento: {item['vencimento']}\n"
                f"Forma: {item['forma_pagamento']}\n"
                f"Status: {item['status']}\n"
                f"Descrição: {item['descricao']}\n"
                f"Valor: {format_currency(float(item['valor']))}"
            )
            print(f"[Telegram] Resposta enviada ao chat_id={chat_id}")
            return

        if acao["tipo"] == "saldo":
            print("[Telegram] Processando comando saldo")
            resumo = calcular_resumo()
            print(f"[Telegram] Resumo calculado: {resumo}")

            telegram_send_message(
                chat_id,
                "💰 Saldo atual\n\n"
                f"Entradas: {format_currency(resumo['total_entradas'])}\n"
                f"Saídas: {format_currency(resumo['total_saidas'])}\n"
                f"Cartão de Crédito: {format_currency(resumo['total_cartao_credito'])}\n"
                f"Saldo Atual: {format_currency(resumo['saldo_atual'])}"
            )
            print(f"[Telegram] Resposta enviada ao chat_id={chat_id}")
            return

        if acao["tipo"] == "resumo":
            print("[Telegram] Processando comando resumo")
            resumo = calcular_resumo()
            print(f"[Telegram] Resumo calculado: {resumo}")

            telegram_send_message(
                chat_id,
                "📊 Resumo financeiro\n\n"
                f"Entradas: {format_currency(resumo['total_entradas'])}\n"
                f"Saídas: {format_currency(resumo['total_saidas'])}\n"
                f"Cartão de Crédito: {format_currency(resumo['total_cartao_credito'])}\n"
                f"Saldo Atual: {format_currency(resumo['saldo_atual'])}"
            )
            print(f"[Telegram] Resposta enviada ao chat_id={chat_id}")
            return

        telegram_send_message(
            chat_id,
            "Comandos disponíveis:\n\n"
            "1) entrada 2500 salario Recebimento cliente\n"
            "2) saida 89,90 pix internet\n"
            "3) saída 350 cartao de credito combustivel\n"
            "4) despesa 1200 boleto aluguel 20/04/2026\n"
            "5) despesa 3000 pix cartao nubank camila pago\n"
            "6) saldo\n"
            "7) resumo"
        )
        print("[Telegram] Ajuda enviada")

    except Exception as exc:
        print(f"[Telegram] Erro no processamento: {exc}")
        telegram_send_message(chat_id, f"❌ Erro ao processar mensagem: {exc}")


def telegram_polling_loop():
    global telegram_offset

    print("[Telegram] Polling iniciado com getUpdates.")

    while True:
        try:
            updates = telegram_get_updates(telegram_offset)

            for update in updates:
                telegram_offset = update["update_id"] + 1
                processar_update(update)

        except Exception as exc:
            print(f"[Telegram] Falha no loop principal: {exc}")

        time.sleep(2)


@app.on_event("startup")
def startup_event():
    init_workbook()
    print(f"[API] XLSX_PATH = {XLSX_PATH}")
    print(f"[API] ALLOWED_ORIGINS = {ALLOWED_ORIGINS}")

    if TELEGRAM_BOT_TOKEN:
        telegram_get_me()
        telegram_delete_webhook()

        t = Thread(target=telegram_polling_loop, daemon=True)
        t.start()
    else:
        print("[Telegram] Token não configurado.")


@app.get("/api/health")
def health_check():
    return {
        "status": "ok",
        "xlsx": str(XLSX_PATH),
        "xlsx_exists": XLSX_PATH.exists(),
        "telegram": "ativo" if TELEGRAM_BOT_TOKEN else "desativado",
        "allowed_chat_id": TELEGRAM_ALLOWED_CHAT_ID or "não definido",
        "allowed_origins": ALLOWED_ORIGINS,
    }


@app.get("/")
def root():
    return {
        "message": "Fluxo de Caixa API online",
        "docs": "/docs",
        "health": "/api/health",
    }


@app.get("/api/entradas", response_model=list[EntradaOut])
def list_entradas():
    return rows_as_dicts("Entradas")


@app.post("/api/entradas", response_model=EntradaOut)
def create_entrada(payload: EntradaIn):
    return append_row("Entradas", payload.model_dump())


@app.put("/api/entradas/{item_id}", response_model=EntradaOut)
def put_entrada(item_id: int, payload: EntradaIn):
    return update_row("Entradas", item_id, payload.model_dump())


@app.delete("/api/entradas/{item_id}")
def remove_entrada(item_id: int):
    return delete_row("Entradas", item_id)


@app.get("/api/saidas", response_model=list[SaidaOut])
def list_saidas():
    return rows_as_dicts("Saidas")


@app.post("/api/saidas", response_model=SaidaOut)
def create_saida(payload: SaidaIn):
    return append_row("Saidas", payload.model_dump())


@app.put("/api/saidas/{item_id}", response_model=SaidaOut)
def put_saida(item_id: int, payload: SaidaIn):
    return update_row("Saidas", item_id, payload.model_dump())


@app.delete("/api/saidas/{item_id}")
def remove_saida(item_id: int):
    return delete_row("Saidas", item_id)


@app.get("/api/despesas", response_model=list[DespesaOut])
def list_despesas():
    return rows_as_dicts("DespesasMes")


@app.post("/api/despesas", response_model=DespesaOut)
def create_despesa(payload: DespesaIn):
    return append_row("DespesasMes", payload.model_dump())


@app.put("/api/despesas/{item_id}", response_model=DespesaOut)
def put_despesa(item_id: int, payload: DespesaIn):
    return update_row("DespesasMes", item_id, payload.model_dump())


@app.delete("/api/despesas/{item_id}")
def remove_despesa(item_id: int):
    return delete_row("DespesasMes", item_id)